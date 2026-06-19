#!/usr/bin/env python3
"""
build_pra.py — Genera PRA .xlsx patchando il template SIAE.
Uso: python3 build_pra.py <mtp_data.json> [output_dir]

Patcha: Copertina (processo), Informazioni (anagrafica),
        Obiettivi del Test (dati rischio R5+), Piano (sprint R2+).
NON tocca: Matrice Rischio, Tabelle.
"""

import sys
import json
import shutil
import traceback
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRORE: openpyxl non installato. Esegui: pip3 install openpyxl")
    sys.exit(1)

SKILL_DIR     = Path(__file__).parent.parent
TEMPLATE_PATH = SKILL_DIR / "assets" / "PRA_template.xlsx"

RISCHIO_FORMULA_TEMPLATE = (
    '=IF(B{r}="Bassa",IF(E{r}="Bassa","Bassa",IF(E{r}="Media","Bassa",'
    'IF(E{r}="Alta","Media",IF(E{r}="Molto Alta","Alta","")))),IF(B{r}="Media",'
    'IF(E{r}="Bassa","Bassa",IF(E{r}="Media","Media",IF(E{r}="Alta","Media",'
    'IF(E{r}="Molto Alta","Alta",0)))),IF(B{r}="Alta",IF(E{r}="Bassa","Media",'
    'IF(E{r}="Media","Alta",IF(E{r}="Alta","Alta",IF(E{r}="Molto Alta","Molto Alta","")))),IF(B{r}="Bloccante",'
    'IF(E{r}="Bassa","Alta",IF(E{r}="Media","Alta",IF(E{r}="Alta","Molto Alta",'
    'IF(E{r}="Molto Alta","Molto Alta",""))))))))'
)


def load_data(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def rischio_formula(row_num):
    return RISCHIO_FORMULA_TEMPLATE.format(r=row_num)


def patch_pra(json_path, output_dir):
    data = load_data(json_path)
    meta = data.get("meta", {})
    codice   = meta.get("codice", "PRA")
    out_name = f"PRA_-_{codice}.xlsx"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / out_name

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template PRA non trovato: {TEMPLATE_PATH}\n"
            f"Verifica che assets/PRA_template.xlsx sia presente."
        )

    shutil.copy2(str(TEMPLATE_PATH), str(output_path))
    print(f"Template: {TEMPLATE_PATH}")
    print(f"Output:   {output_path}")

    wb = openpyxl.load_workbook(str(output_path))

    # ----------------------------------------------------------------
    # Copertina — R10 F = titolo/processo
    # ----------------------------------------------------------------
    ws_cov = wb["Copertina"]
    ws_cov.cell(10, 6).value = meta.get("titolo", "")
    print(f"  Copertina: processo = '{meta.get('titolo', '')}'")

    # ----------------------------------------------------------------
    # Informazioni — R6: versione, data, autore, descrizione
    # ----------------------------------------------------------------
    ws_inf = wb["Informazioni"]
    ws_inf.cell(6, 3).value = meta.get("versione", "1.0")
    ws_inf.cell(6, 4).value = meta.get("data", "")
    ws_inf.cell(6, 5).value = meta.get("autore", "")
    ws_inf.cell(6, 6).value = ""                 # validatore (opzionale)
    ws_inf.cell(6, 7).value = "Prima Emissione"
    print(f"  Informazioni: v{meta.get('versione','1.0')}, {meta.get('data','')}, {meta.get('autore','')}")

    # ----------------------------------------------------------------
    # Obiettivi del Test — dati da R5 in poi
    # ----------------------------------------------------------------
    ws_obj = wb["Obiettivi del Test"]
    obiettivi = data.get("pra_obiettivi", [])

    # Svuota righe dati esistenti (R5+) — colonne A-H (non toccare I con formula)
    for row in ws_obj.iter_rows(min_row=5, max_row=ws_obj.max_row):
        for cell in row:
            cell.value = None   # svuota tutto incluse le vecchie formule

    # Scrivi nuovi dati + formula rischio
    for i, obj in enumerate(obiettivi):
        r = 5 + i
        ws_obj.cell(r, 1).value = obj.get("obiettivo", "")
        ws_obj.cell(r, 2).value = obj.get("criticita", "Alta")
        ws_obj.cell(r, 3).value = obj.get("fattore_rischio", "Q-15 Aderenza ai requisiti")
        ws_obj.cell(r, 4).value = obj.get("motivazione", "")
        ws_obj.cell(r, 5).value = obj.get("frequenza_uso", "Alta")
        ws_obj.cell(r, 6).value = obj.get("req_nrt", "No")
        ws_obj.cell(r, 7).value = obj.get("req_performance", "No")
        ws_obj.cell(r, 8).value = obj.get("req_e2e_uat", "Si")
        ws_obj.cell(r, 9).value = rischio_formula(r)

    print(f"  Obiettivi del Test: {len(obiettivi)} righe scritte (R5-R{4+len(obiettivi)})")

    # ----------------------------------------------------------------
    # Piano — dati da R2 in poi
    # ----------------------------------------------------------------
    ws_piano = wb["Piano"]
    piano = data.get("pra_piano", [])

    # Svuota righe dati esistenti (R2+)
    for row in ws_piano.iter_rows(min_row=2, max_row=ws_piano.max_row):
        for cell in row:
            cell.value = None

    for i, entry in enumerate(piano):
        r = 2 + i
        ws_piano.cell(r, 1).value = entry.get("team", "")
        ws_piano.cell(r, 2).value = entry.get("iterazione", "")
        ws_piano.cell(r, 3).value = entry.get("processo", meta.get("titolo", ""))
        ws_piano.cell(r, 4).value = entry.get("owner", "")

    print(f"  Piano: {len(piano)} righe scritte")

    wb.save(str(output_path))
    print(f"\nGenerato: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Uso: {sys.argv[0]} <mtp_data.json> [output_dir]")
        sys.exit(1)

    json_path = sys.argv[1]
    out_dir   = sys.argv[2] if len(sys.argv) > 2 else "."

    try:
        result = patch_pra(json_path, out_dir)
        print(f"OK: {result}")
    except Exception as e:
        print(f"ERRORE: {e}")
        traceback.print_exc()
        sys.exit(1)
