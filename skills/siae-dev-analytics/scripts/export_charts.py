"""openpyxl charts factory per Excel report v2.

Chart types:
- Velocity bar chart per developer
- Velocity x Quality scatter (quadrant analysis)
- Before/After grouped bar (AI Impact)
- PR types pie chart
- AI-assisted vs Manual pie chart
"""
from __future__ import annotations

import logging

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series

log = logging.getLogger(__name__)


def add_velocity_bar_chart(ws, data_sheet_name: str, n_devs: int) -> None:
    """Bar chart indice velocita per dev. Dati da sheet 'Per Developer'."""
    if n_devs < 1:
        log.info("add_velocity_bar_chart: skip — no dev data")
        return
    chart = BarChart()
    chart.title = "Indice Velocita per Sviluppatore"
    chart.y_axis.title = "Indice velocita (z-score)"
    chart.x_axis.title = "Sviluppatore"
    data = Reference(ws.parent[data_sheet_name], min_col=13, min_row=1, max_row=n_devs + 1, max_col=13)
    cats = Reference(ws.parent[data_sheet_name], min_col=1, min_row=2, max_row=n_devs + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, "F3")


def add_velocity_quality_scatter(ws, data_sheet_name: str, n_devs: int) -> None:
    """Scatter velocity x quality (quadrant analysis)."""
    if n_devs < 1:
        log.info("add_velocity_quality_scatter: skip — no dev data")
        return
    chart = ScatterChart()
    chart.title = "Velocita x Qualita (quadrant analysis)"
    chart.style = 13
    chart.x_axis.title = "Velocita"
    chart.y_axis.title = "Qualita"
    x = Reference(ws.parent[data_sheet_name], min_col=13, min_row=2, max_row=n_devs + 1)
    y = Reference(ws.parent[data_sheet_name], min_col=14, min_row=2, max_row=n_devs + 1)
    series = Series(y, x, title_from_data=False)
    chart.series.append(series)
    chart.height = 10
    chart.width = 15
    ws.add_chart(chart, "F20")


def add_before_after_bar_chart(
    ws, kpi_names: list[str], baseline_vals: list[float], current_vals: list[float]
) -> None:
    """Grouped bar chart: baseline vs current per KPI."""
    if not kpi_names:
        log.info("add_before_after_bar_chart: skip — empty kpi_names")
        return
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "AI Impact - Before / After"
    chart.y_axis.title = "Valore"
    start_row = 50
    ws.cell(row=start_row, column=1, value="KPI")
    ws.cell(row=start_row, column=2, value="Baseline")
    ws.cell(row=start_row, column=3, value="Current")
    for i, (name, b, c) in enumerate(zip(kpi_names, baseline_vals, current_vals), start=start_row + 1):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(kpi_names), max_col=3)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(kpi_names))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E50")


def add_pr_types_pie_chart(ws, types: dict[str, int]) -> None:
    """Pie chart distribuzione PR per tipo."""
    if not types:
        log.info("add_pr_types_pie_chart: skip — empty types")
        return
    chart = PieChart()
    chart.title = "Distribuzione PR per Tipo (feat/fix/refactor/...)"
    start_row = 80
    for i, (t, c) in enumerate(types.items(), start=start_row):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=c)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(types) - 1)
    labels = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(types) - 1)
    chart.add_data(data)
    chart.set_categories(labels)
    ws.add_chart(chart, "E80")


def add_ai_vs_manual_pie(ws, ai_pct: float) -> None:
    """Pie chart AI-assisted vs Manual."""
    chart = PieChart()
    chart.title = "PR AI-assisted vs Manuali"
    ws.cell(row=100, column=1, value="AI-assisted")
    ws.cell(row=100, column=2, value=ai_pct)
    ws.cell(row=101, column=1, value="Manual")
    ws.cell(row=101, column=2, value=max(0, 1 - ai_pct))
    data = Reference(ws, min_col=2, min_row=100, max_row=101)
    labels = Reference(ws, min_col=1, min_row=100, max_row=101)
    chart.add_data(data)
    chart.set_categories(labels)
    ws.add_chart(chart, "E100")
