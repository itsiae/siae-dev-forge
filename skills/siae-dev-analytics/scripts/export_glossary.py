"""Sheet 'Glossario KPI' per Executive Excel report."""
from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_glossary_data(yaml_path: Path) -> list[dict]:
    """Carica glossario da YAML. Empty list se file mancante."""
    if not yaml_path.exists():
        return []
    try:
        data = yaml.safe_load(yaml_path.read_text())
        return data.get("kpi", []) if data else []
    except Exception:
        return []


def write_glossary_sheet(wb: Workbook, glossary: list[dict]) -> None:
    """Scrive sheet 'Glossario KPI' come tab secondario verde."""
    ws = wb.create_sheet("Glossario KPI")
    ws.sheet_properties.tabColor = "63BE7B"  # verde

    bold = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="E8F5E9")

    ws["A1"] = "Glossario KPI - DevForge Dev Analytics"
    ws["A1"].font = Font(bold=True, size=14, color="0D6E3D")
    ws.merge_cells("A1:E1")

    headers = ["KPI", "Nome italiano", "Formula", "Interpretazione", "Buono se", "Benchmark industria"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill

    for i, kpi in enumerate(glossary, start=4):
        ws.cell(row=i, column=1, value=kpi.get("id", ""))
        ws.cell(row=i, column=2, value=kpi.get("label_it", ""))
        ws.cell(row=i, column=3, value=kpi.get("formula", ""))
        ws.cell(row=i, column=4, value=kpi.get("interpretation", ""))
        ws.cell(row=i, column=5, value=kpi.get("good_if", ""))
        ws.cell(row=i, column=6, value=kpi.get("benchmark", ""))
        ws.row_dimensions[i].height = 40
        for c in range(1, 7):
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [25, 30, 35, 45, 20, 30]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A4"
