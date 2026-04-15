"""Export KPI report in Excel xlsx con 5 sheet (executive-friendly).

Sheets (in ordine):
    1. Executive Summary — health score, DORA tier, narrative, data quality
    2. Summary — metadata + Top/Bottom 5 ROI (audit-friendly)
    3. Per Developer — 11 KPI + score (colonne italiane)
    4. Raw Data — denormalized PR dump
    5. Data Sources — mode + KPI calculability
"""
from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# Mapping colonne tecniche → label executive italiani
EXEC_COLUMN_LABELS = {
    "pr_cycle_time_p50": "Tempo mediano PR (h)",
    "lead_time_to_merge_p50": "Lead time mediano (h)",
    "pr_throughput_weekly": "PR per settimana",
    "time_to_first_review_p50": "Tempo mediano prima review (h)",
    "deploy_frequency_monthly": "Deploy al mese",
    "review_comments_p50": "Commenti review (mediana)",
    "rework_ratio": "% PR con rework",
    "test_presence_rate": "% PR con test",
    "verification_rate": "% commit verificati",
    "design_driven_rate": "% PR con design doc",
    "revert_rate": "% commit di revert",
    "velocity_score": "Indice velocità",
    "quality_score": "Indice qualità",
    "roi_index": "Indice ROI",
}


def _anonymize_index(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.index = [hashlib.sha256(str(i).encode()).hexdigest()[:8] for i in df.index]
    return df


def _anonymize_column(s: pd.Series) -> pd.Series:
    return s.apply(lambda v: hashlib.sha256(str(v).encode()).hexdigest()[:8])


def _dora_tier_from_cycle_time(median_cycle_hours: float | None) -> tuple[str, str]:
    """DORA Lead Time tier da median cycle time. Returns (tier, hex_color).

    Boundaries (DORA 2023 State of DevOps):
        [0, 24h)    Elite
        [24h, 168h) High      (< 1 settimana)
        [168h, 720h) Medium   (< 1 mese)
        [720h, inf)  Low      (>= 1 mese)

    None/NaN → N/A (dato mancante — uno 0.0 reale resta Elite).
    """
    if median_cycle_hours is None or pd.isna(median_cycle_hours):
        return ("N/A", "CCCCCC")
    if median_cycle_hours < 24:
        return ("Elite", "63BE7B")
    if median_cycle_hours < 168:
        return ("High", "A4C96E")
    if median_cycle_hours < 720:
        return ("Medium", "FFEB9C")
    return ("Low", "FF6B6B")


# Minimo dev per z-score statisticamente significativo
HEALTH_SCORE_MIN_SAMPLE = 3


def _team_health_score(kpis_df: pd.DataFrame) -> tuple[int, str, str]:
    """Team health 0-100 + tier + hex_color.

    Contratto: velocity_score e quality_score sono z-score calcolati da
    compute_kpis. Con N<3 dev il campione è troppo piccolo per z-score
    significativi → restituiamo tier "N/A - campione insufficiente".

    Mapping: z ∈ [-3, +3] → [0, 100] linear, clamp agli estremi.
    """
    if kpis_df.empty or "velocity_score" not in kpis_df.columns:
        return (0, "N/A", "CCCCCC")
    if len(kpis_df) < HEALTH_SCORE_MIN_SAMPLE:
        return (0, "N/A - campione insufficiente", "CCCCCC")
    avg_vel = kpis_df["velocity_score"].mean()
    avg_qual = kpis_df["quality_score"].mean()
    vel_norm = max(0, min(100, (avg_vel + 3) * 100 / 6))
    qual_norm = max(0, min(100, (avg_qual + 3) * 100 / 6))
    score = int(round((vel_norm + qual_norm) / 2))
    if score >= 75:
        return (score, "Eccellente 🟢", "63BE7B")
    if score >= 55:
        return (score, "Buono 🟢", "A4C96E")
    if score >= 40:
        return (score, "Adeguato 🟡", "FFEB9C")
    if score >= 25:
        return (score, "Da migliorare 🟠", "FFA500")
    return (score, "Critico 🔴", "FF6B6B")


def _data_quality_warnings(kpis_df: pd.DataFrame) -> list[str]:
    """Avvisi qualità dati per KPI che sono 0 per tutti."""
    warnings = []
    if kpis_df.empty:
        return ["Dataset vuoto — nessun dato nel periodo."]
    checks = [
        ("deploy_frequency_monthly", "Deploy frequency = 0: repo non usa tag convention SIAE (COLLAUDO/CERT/PROD) → KPI N/A."),
        ("verification_rate", "Verification rate = 0: trailer `verified-by: siae-verification` non adottato → KPI non misurabile."),
        ("design_driven_rate", "Design-driven rate = 0: PR non linkano design doc → adozione `siae-brainstorming` non tracciabile."),
        ("rework_ratio", "Rework ratio = 0: KPI Q2 deferred in v1 (richiede timeline API)."),
    ]
    for col, msg in checks:
        if col in kpis_df.columns and (kpis_df[col] == 0).all():
            warnings.append(msg)
    return warnings


def _executive_insights(kpis_df: pd.DataFrame, n_prs: int) -> list[str]:
    """3-5 bullet narrative per executive in italiano."""
    insights = []
    if kpis_df.empty:
        return ["Nessun dato sufficiente per insight."]

    if "pr_cycle_time_p50" in kpis_df:
        median_cycle = kpis_df["pr_cycle_time_p50"].median()
        tier, _ = _dora_tier_from_cycle_time(median_cycle)
        insights.append(f"Il team chiude una PR in mediana {median_cycle:.1f} ore — classificazione DORA: {tier}.")

    if "pr_throughput_weekly" in kpis_df:
        total = kpis_df["pr_throughput_weekly"].sum()
        insights.append(f"Throughput aggregato: {total:.1f} PR/settimana su {len(kpis_df)} dev ({n_prs} PR totali).")

    if "test_presence_rate" in kpis_df:
        rate = kpis_df["test_presence_rate"].mean()
        if rate >= 0.7:
            insights.append(f"Qualità test buona: {rate*100:.0f}% PR include test.")
        elif rate >= 0.3:
            insights.append(f"Copertura test moderata: {rate*100:.0f}% PR con test. Target consigliato ≥ 70%.")
        else:
            insights.append(f"⚠️ Copertura test bassa: {rate*100:.0f}% PR con test — rischio regressioni elevato.")

    if "design_driven_rate" in kpis_df:
        rate = kpis_df["design_driven_rate"].mean()
        if rate >= 0.3:
            insights.append(f"Design-driven development: {rate*100:.0f}% PR linkano design doc.")
        elif rate > 0:
            insights.append(f"Design-driven development poco adottato ({rate*100:.0f}%). Opportunità: promuovere `siae-brainstorming`.")

    return insights


def _write_executive_summary_sheet(wb: Workbook, kpis_df: pd.DataFrame, source_report: dict,
                                     window: tuple[str, str], generated_at: str, n_prs: int) -> None:
    """Scrive sheet 'Executive Summary' come primo tab."""
    ws = wb.create_sheet("Executive Summary", 0)  # posizione 0 = primo tab
    bold = Font(bold=True, size=12)
    big = Font(bold=True, size=16)
    title_fill = PatternFill("solid", fgColor="FFF3CD")

    ws["A1"] = "CONFIDENZIALE — Dati Personali SIAE"
    ws["A1"].font = Font(bold=True, color="CC0000", size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:D1")

    ws["A3"] = "Executive Summary — DevForge Analytics"
    ws["A3"].font = Font(bold=True, size=16)
    ws.merge_cells("A3:D3")

    ws["A5"] = "Periodo:"; ws["B5"] = f"{window[0]} → {window[1]}"
    ws["A6"] = "Sviluppatori:"; ws["B6"] = len(kpis_df)
    ws["A7"] = "PR totali:"; ws["B7"] = n_prs
    ws["A8"] = "Report generato:"; ws["B8"] = generated_at

    # Team Health Score
    score, tier, hex_color = _team_health_score(kpis_df)
    ws["A10"] = "Team Health Score"; ws["A10"].font = bold
    ws["A11"] = score; ws["A11"].font = big
    ws["A11"].fill = PatternFill("solid", fgColor=hex_color)
    ws["B11"] = f"/ 100 — {tier}"; ws["B11"].font = bold
    ws.merge_cells("B11:D11")

    # DORA tier
    if not kpis_df.empty and "pr_cycle_time_p50" in kpis_df.columns:
        median_cycle = kpis_df["pr_cycle_time_p50"].median()
    else:
        median_cycle = None
    dora_tier, dora_color = _dora_tier_from_cycle_time(median_cycle)
    ws["A13"] = "DORA TIER (cycle time)"; ws["A13"].font = bold
    ws["A14"] = dora_tier; ws["A14"].font = big
    ws["A14"].fill = PatternFill("solid", fgColor=dora_color)
    ws["B14"] = f"Mediana: {median_cycle:.1f} ore" if median_cycle is not None else "Mediana: N/A"
    ws.merge_cells("B14:D14")

    # Top 3 performers (se disponibili)
    row = 16
    if not kpis_df.empty and "roi_index" in kpis_df.columns:
        top3 = kpis_df.sort_values("roi_index", ascending=False).head(3)
        ws[f"A{row}"] = "🏆 TOP 3 PERFORMER (per ROI)"; ws[f"A{row}"].font = bold
        row += 1
        ws[f"A{row}"] = "Dev"; ws[f"B{row}"] = "Indice ROI"; ws[f"C{row}"] = "Velocità"; ws[f"D{row}"] = "Qualità"
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = bold
        row += 1
        for dev, r in top3.iterrows():
            ws[f"A{row}"] = str(dev)
            ws[f"B{row}"] = round(r["roi_index"], 3) if "roi_index" in r else 0
            ws[f"C{row}"] = round(r["velocity_score"], 3) if "velocity_score" in r else 0
            ws[f"D{row}"] = round(r["quality_score"], 3) if "quality_score" in r else 0
            row += 1
        row += 1

        # Attention needed (bottom 3)
        bot3 = kpis_df.sort_values("roi_index", ascending=True).head(3)
        ws[f"A{row}"] = "⚠️ ATTENZIONE (score più bassi)"; ws[f"A{row}"].font = bold
        row += 1
        ws[f"A{row}"] = "Dev"; ws[f"B{row}"] = "Indice ROI"; ws[f"C{row}"] = "Velocità"; ws[f"D{row}"] = "Qualità"
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = bold
        row += 1
        for dev, r in bot3.iterrows():
            ws[f"A{row}"] = str(dev)
            ws[f"B{row}"] = round(r["roi_index"], 3) if "roi_index" in r else 0
            ws[f"C{row}"] = round(r["velocity_score"], 3) if "velocity_score" in r else 0
            ws[f"D{row}"] = round(r["quality_score"], 3) if "quality_score" in r else 0
            row += 1
        row += 1

    # Insights narrative
    insights = _executive_insights(kpis_df, n_prs)
    ws[f"A{row}"] = "💡 INSIGHT"; ws[f"A{row}"].font = bold
    row += 1
    for ins in insights:
        ws[f"A{row}"] = f"• {ins}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 30
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1

    # Data quality warnings
    warnings = _data_quality_warnings(kpis_df)
    if warnings:
        ws[f"A{row}"] = "⚠️ QUALITÀ DATI (KPI N/A per questo dataset)"; ws[f"A{row}"].font = bold
        row += 1
        for w in warnings:
            ws[f"A{row}"] = f"• {w}"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.row_dimensions[row].height = 30
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    for col_idx in range(1, 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28


def export(
    kpis_df: pd.DataFrame,
    raw_prs: pd.DataFrame,
    source_report: dict,
    window: tuple[str, str],
    output_path: Path,
    *,
    anonymize: bool = False,
    generated_at: str | None = None,
) -> Path:
    """Export DataFrame KPI in xlsx 4-sheet."""
    generated_at = generated_at or datetime.utcnow().isoformat() + "Z"

    if anonymize:
        kpis_df = _anonymize_index(kpis_df)
        if "author" in raw_prs.columns:
            raw_prs = raw_prs.copy()
            raw_prs["author"] = _anonymize_column(raw_prs["author"])

    wb = Workbook()

    # -- Sheet 1: Summary -------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    bold = Font(bold=True, size=12)
    title_fill = PatternFill("solid", fgColor="FFF3CD")

    ws["A1"] = "CONFIDENZIALE — Dati Personali SIAE"
    ws["A1"].font = Font(bold=True, color="CC0000", size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:D1")

    ws["A3"] = "DevForge Dev Analytics Report"
    ws["A3"].font = bold
    ws["A4"] = "Generated at:"; ws["B4"] = generated_at
    ws["A5"] = "Mode:"; ws["B5"] = source_report.get("mode", "UNKNOWN")
    ws["A6"] = "Window:"; ws["B6"] = f"{window[0]} → {window[1]}"
    ws["A7"] = "Developers:"; ws["B7"] = len(kpis_df)
    ws["A8"] = "Anonymized:"; ws["B8"] = "YES" if anonymize else "NO"

    # Top/Bottom 5 ROI
    ws["A10"] = "Top 5 ROI Index"; ws["A10"].font = bold
    top5 = kpis_df.sort_values("roi_index", ascending=False).head(5)
    ws["A11"] = "Dev"; ws["B11"] = "ROI"; ws["C11"] = "Velocity"; ws["D11"] = "Quality"
    for i, (dev, row) in enumerate(top5.iterrows(), start=12):
        ws[f"A{i}"] = dev
        ws[f"B{i}"] = round(row["roi_index"], 3)
        ws[f"C{i}"] = round(row["velocity_score"], 3)
        ws[f"D{i}"] = round(row["quality_score"], 3)

    bottom_start = 12 + len(top5) + 2
    ws[f"A{bottom_start}"] = "Bottom 5 ROI Index"; ws[f"A{bottom_start}"].font = bold
    bottom5 = kpis_df.sort_values("roi_index", ascending=True).head(5)
    ws[f"A{bottom_start+1}"] = "Dev"; ws[f"B{bottom_start+1}"] = "ROI"
    ws[f"C{bottom_start+1}"] = "Velocity"; ws[f"D{bottom_start+1}"] = "Quality"
    for i, (dev, row) in enumerate(bottom5.iterrows(), start=bottom_start + 2):
        ws[f"A{i}"] = dev
        ws[f"B{i}"] = round(row["roi_index"], 3)
        ws[f"C{i}"] = round(row["velocity_score"], 3)
        ws[f"D{i}"] = round(row["quality_score"], 3)

    for col_idx in range(1, 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # -- Sheet Executive Summary (inserito come primo tab) -----------------
    n_prs = len(raw_prs) if not raw_prs.empty else 0
    _write_executive_summary_sheet(wb, kpis_df, source_report, window, generated_at, n_prs)

    # -- Sheet Per Developer (header italiani) ----------------------------
    ws2 = wb.create_sheet("Per Developer")
    cols = ["dev"] + list(kpis_df.columns)
    # Header: usa label italiani exec-friendly se mappati, altrimenti tecnico
    header_labels = ["Sviluppatore"] + [EXEC_COLUMN_LABELS.get(c, c) for c in kpis_df.columns]
    ws2.append(header_labels)
    for c in range(1, len(cols) + 1):
        ws2.cell(row=1, column=c).font = bold

    for dev, row in kpis_df.iterrows():
        ws2.append([dev] + [round(v, 3) if isinstance(v, (int, float)) else v for v in row])

    # Conditional formatting su z-score columns (velocity_score, quality_score, roi_index)
    n_rows = len(kpis_df) + 1
    for score_col in ["velocity_score", "quality_score", "roi_index"]:
        if score_col in cols:
            c_idx = cols.index(score_col) + 1
            col_letter = get_column_letter(c_idx)
            rng = f"{col_letter}2:{col_letter}{n_rows}"
            rule = ColorScaleRule(
                start_type="min", start_color="FF6B6B",
                mid_type="num", mid_value=0, mid_color="FFEB9C",
                end_type="max", end_color="63BE7B",
            )
            ws2.conditional_formatting.add(rng, rule)

    ws2.freeze_panes = "B2"
    for col_idx in range(1, len(cols) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # -- Sheet 3: Raw Data ------------------------------------------------
    ws3 = wb.create_sheet("Raw Data")
    if not raw_prs.empty:
        ws3.append(list(raw_prs.columns))
        for c in range(1, len(raw_prs.columns) + 1):
            ws3.cell(row=1, column=c).font = bold
        for _, row in raw_prs.iterrows():
            ws3.append([str(v) if v is not None else "" for v in row])
    else:
        ws3["A1"] = "No raw data available"

    # -- Sheet 4: Data Sources --------------------------------------------
    ws4 = wb.create_sheet("Data Sources")
    ws4["A1"] = "Data Sources Declaration"
    ws4["A1"].font = Font(bold=True, size=14)
    ws4["A3"] = "Mode:"; ws4["B3"] = source_report.get("mode", "UNKNOWN")
    ws4["A4"] = "Generated at:"; ws4["B4"] = generated_at

    ws4["A6"] = "Source"; ws4["B6"] = "Available"; ws4["C6"] = "Notes"
    for c in range(1, 4):
        ws4.cell(row=6, column=c).font = bold

    ws4.append(["GitHub", "YES" if source_report.get("github") else "NO",
                "Ground truth — gh graphql"])
    ws4.append(["S3 devforge-logs", "YES" if source_report.get("s3_devforge") else "NO",
                "DevForge telemetry events"])
    ws4.append(["S3 blend-usage", "YES" if source_report.get("s3_blend") else "NO",
                "Bedrock+Anthropic token blend"])

    ws4["A12"] = "KPI calculability per mode:"; ws4["A12"].font = bold
    mode = source_report.get("mode", "UNKNOWN")
    if mode == "FULL":
        ws4["A13"] = "11/11 KPI + ROI with real cost_score"
    elif mode == "HYBRID":
        ws4["A13"] = "11/11 KPI, ROI = velocity × quality (cost=1)"
    elif mode == "GITHUB-ONLY":
        ws4["A13"] = "11/11 KPI, Q4 best-effort (commit trailer), ROI = velocity × quality"
    else:
        ws4["A13"] = f"Mode {mode} — check configuration"

    for col_idx in range(1, 4):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 25

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import argparse, json
    parser = argparse.ArgumentParser()
    parser.add_argument("--kpis", required=True, help="CSV path con KPI")
    parser.add_argument("--prs", required=True, help="CSV path con PR raw")
    parser.add_argument("--sources", required=True, help="JSON path con source report")
    parser.add_argument("--from", dest="window_from", required=True)
    parser.add_argument("--to", dest="window_to", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--anonymize", action="store_true")
    args = parser.parse_args()

    kpis = pd.read_csv(args.kpis, index_col=0)
    prs = pd.read_csv(args.prs)
    sources = json.loads(Path(args.sources).read_text())
    out = export(kpis, prs, sources, (args.window_from, args.window_to),
                 Path(args.out), anonymize=args.anonymize)
    print(f"Report saved to: {out}")
