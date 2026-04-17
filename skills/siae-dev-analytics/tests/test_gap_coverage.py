"""Gap coverage tests — edge case per raggiungere target 272 test (34 test)."""
from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
import pytest

import autodetect_sources as ad
import collect_anthropic_api as ca
import compute_ai_impact as ai
import compute_branches as cb
import compute_kpis as ck
import compute_reviews as cr
import export_glossary as eg
import seasonality as s
import validators as v


# ── Empty input edge cases (ogni KPI) ──

def test_kpi_open_prs_empty():
    assert ck.kpi_open_prs_count(pd.DataFrame()) == {}

def test_kpi_draft_prs_empty():
    assert ck.kpi_draft_prs_count(pd.DataFrame()) == {}

def test_kpi_stuck_prs_empty():
    assert ck.kpi_stuck_prs_count(pd.DataFrame()) == {}

def test_kpi_closed_unmerged_empty():
    assert ck.kpi_closed_unmerged_count(pd.DataFrame()) == {}

def test_kpi_reopen_empty():
    assert ck.kpi_reopen_count(pd.DataFrame()) == {}

def test_kpi_oldest_open_empty():
    assert ck.kpi_oldest_open_pr_age_days(pd.DataFrame()) == {}

def test_kpi_active_branches_empty():
    assert cb.kpi_active_branches_per_dev(pd.DataFrame()) == {}

def test_kpi_branches_without_pr_empty():
    assert cb.kpi_branches_without_pr(pd.DataFrame()) == {}

def test_kpi_stale_branches_empty():
    assert cb.kpi_stale_branches_count(pd.DataFrame()) == 0

def test_kpi_naming_compliance_empty():
    assert cb.kpi_branch_naming_compliance_rate(pd.DataFrame()) == 0.0

def test_kpi_hotfix_count_empty():
    assert cb.kpi_hotfix_branches_count(pd.DataFrame()) == 0

def test_kpi_reviews_given_empty():
    assert cr.kpi_reviews_given_count(pd.DataFrame()) == {}

def test_kpi_approvals_empty():
    assert cr.kpi_approvals_given_count(pd.DataFrame()) == {}

def test_kpi_co_authored_empty():
    assert cr.kpi_co_authored_prs_count(pd.DataFrame()) == {}

def test_kpi_features_shipped_empty():
    assert ck.kpi_features_shipped(pd.DataFrame()) == {}

def test_kpi_net_loc_missing_column():
    prs = pd.DataFrame([{"author": "alice"}])
    assert ck.kpi_net_loc_shipped(prs) == {}


# ── Unicode edge cases ──

def test_unicode_dev_name_z_score():
    devs = {"alice\u00e9": 10.0, "bob\u00fc": 20.0, "\U0001f600emoji": 30.0}
    result = ck.z_score(devs)
    assert "alice\u00e9" in result

def test_unicode_branch_name_compliance():
    branches = pd.DataFrame([{"name": "feat/\u00e9mile-fix", "author": "a", "pr_count": 1}])
    rate = cb.kpi_branch_naming_compliance_rate(branches)
    assert rate == 1.0


# ── Validator edge cases ──

def test_validator_version_1_accepted():
    cfg = v.AnalyticsConfigV2(
        version=1,
        scope=v.ScopeConfigV2(repos=["owner/repo"]),
        time_window=v.TimeWindowSingle(**{"from": "2026-01-01", "to": "2026-03-01"}),
    )
    assert cfg.version == 1

def test_validator_parallel_fetch_boundary():
    opt = v.OptionsConfigV2(parallel_fetch=1)
    assert opt.parallel_fetch == 1
    opt16 = v.OptionsConfigV2(parallel_fetch=16)
    assert opt16.parallel_fetch == 16
    with pytest.raises(Exception):
        v.OptionsConfigV2(parallel_fetch=0)
    with pytest.raises(Exception):
        v.OptionsConfigV2(parallel_fetch=17)

def test_assert_finite_normal():
    v.assert_finite(42.0, "test")  # no raise

def test_assert_finite_nan():
    with pytest.raises(ValueError):
        v.assert_finite(float("nan"), "test")

def test_assert_finite_inf():
    with pytest.raises(ValueError):
        v.assert_finite(float("inf"), "test")


# ── Seasonality edge ──

def test_easter_2025():
    from datetime import date
    assert s.easter_date(2025) == date(2025, 4, 20)

def test_working_days_cross_year():
    wd = s.working_days_in_window("2025-12-29", "2026-01-02")
    assert wd >= 0

def test_seasonality_adj_single_day_weekend():
    adj = s.seasonality_adj("2026-03-07", "2026-03-07")  # Saturday
    assert adj == 0.0


# ── AI Impact edge ──

def test_is_ai_assisted_random_trailer_false():
    assert ai.is_ai_assisted("feat: x\n\nCo-Authored-By: Human <h@h.com>") is False

def test_compute_delta_zero_baseline_zero_current():
    d = ai.compute_delta(0.0, 0.0)
    assert d.trend == "STABLE"
    assert d.delta_pct == 0.0

def test_ai_velocity_multiplier_zero_cycle():
    assert ai.kpi_ai_velocity_multiplier({"a": 0.0}, {"b": 10.0}) == 1.0


# ── Anthropic API edge ──

def test_usd_to_eur_conversion():
    eur = ca.usd_to_eur(10.0)
    assert 8.0 < eur < 10.0  # rate ~0.92


# ── Glossary round-trip ──

def test_glossary_round_trip(tmp_path):
    from openpyxl import Workbook, load_workbook
    ref = Path(__file__).parent.parent / "reference" / "kpi-glossary-data.yaml"
    glossary = eg.load_glossary_data(ref)
    wb = Workbook()
    eg.write_glossary_sheet(wb, glossary)
    out = tmp_path / "roundtrip.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["Glossario KPI"]
    # Prima riga KPI dati (row 4)
    assert ws.cell(row=4, column=1).value is not None
