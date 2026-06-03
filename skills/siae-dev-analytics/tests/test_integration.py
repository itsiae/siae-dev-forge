"""Integration test end-to-end: config → autodetect → fetch → compute → export."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch
import pytest

import run_analytics as ra


@pytest.fixture
def tmp_config(tmp_path):
    cfg = {
        "version": 1,
        "scope": {"repos": ["itsiae/sample-repo"], "teams": [], "topics": []},
        "time_window": {"from": "2026-03-01", "to": "2026-03-31"},
        "developers": {"include": [], "exclude": ["dependabot[bot]"]},
        "options": {"anonymize": False, "min_commits_threshold": 1, "parallel_fetch": 2},
        "output": {"format": "xlsx", "path": str(tmp_path / "report.xlsx")},
    }
    cfg_path = tmp_path / "devforge-analytics.yml"
    import yaml
    cfg_path.write_text(yaml.safe_dump(cfg))
    return cfg_path, cfg


def test_integration_github_only_produces_excel(tmp_config, sample_pr_data, tmp_path, monkeypatch):
    """End-to-end: config → autodetect (GH-ONLY mock) → fetch (mock) → compute → export."""
    cfg_path, cfg = tmp_config

    # Mock autodetect → GITHUB-ONLY
    with patch("autodetect_sources.check_gh_auth", return_value=True), \
         patch("autodetect_sources.check_s3_prefix", return_value=False), \
         patch("collect_github.fetch_repo_data", return_value=sample_pr_data):
        output = ra.cmd_run(cfg_path)

    assert Path(output).exists()
    assert Path(output).suffix == ".xlsx"

    from openpyxl import load_workbook
    wb = load_workbook(output)
    assert set(wb.sheetnames) == {"Executive Summary", "Summary", "Per Developer", "Raw Data", "Data Sources"}


def test_integration_abort_if_no_github(tmp_config):
    """github non autenticato → RuntimeError."""
    cfg_path, _ = tmp_config
    with patch("autodetect_sources.check_gh_auth", return_value=False):
        with pytest.raises(RuntimeError, match="GitHub"):
            ra.cmd_run(cfg_path)


def test_integration_reproducible_checksum(tmp_config, sample_pr_data, tmp_path):
    """Stesso input + generated_at fisso → stesse cell values.

    Passiamo generated_at esplicito via parametro (evita mock datetime fragile)."""
    cfg_path, _ = tmp_config

    from openpyxl import load_workbook

    FIXED_TS = "2026-04-15T00:00:00Z"

    outputs = []
    for suffix in ("a", "b"):
        out_path = tmp_path / f"report_{suffix}.xlsx"
        with patch("autodetect_sources.check_gh_auth", return_value=True), \
             patch("autodetect_sources.check_s3_prefix", return_value=False), \
             patch("collect_github.fetch_repo_data", return_value=sample_pr_data):
            # cmd_run accetta generated_at opzionale per riproducibilità in test
            ra.cmd_run(cfg_path, output_override=out_path, generated_at_override=FIXED_TS)
        outputs.append(out_path)

    wb1 = load_workbook(outputs[0])
    wb2 = load_workbook(outputs[1])
    for sheet in wb1.sheetnames:
        vals1 = [c.value for row in wb1[sheet].iter_rows() for c in row]
        vals2 = [c.value for row in wb2[sheet].iter_rows() for c in row]
        assert vals1 == vals2, f"sheet {sheet} differs"


def test_config_missing_scope_raises(tmp_path):
    """Config senza scope → ValueError."""
    bad_cfg = tmp_path / "bad.yml"
    bad_cfg.write_text("version: 1\n")
    with pytest.raises(ValueError):
        ra.load_config(bad_cfg)


def test_config_empty_scope_raises(tmp_path):
    """Config con scope vuoto → ValueError."""
    bad_cfg = tmp_path / "bad.yml"
    bad_cfg.write_text("""
version: 1
scope:
  repos: []
  teams: []
  topics: []
time_window: {from: '2026-01-01'}
""")
    with pytest.raises(ValueError):
        ra.load_config(bad_cfg)


def test_no_data_produces_txt_not_xlsx(tmp_config, tmp_path):
    """Se fetch ritorna 0 PR e 0 commit → output .no-data.txt."""
    cfg_path, _ = tmp_config
    empty_raw = {
        "repository": {
            "nameWithOwner": "itsiae/empty",
            "pullRequests": {"nodes": []},
            "defaultBranchRef": {"target": {"history": {"nodes": []}}},
            "refs": {"nodes": []},
        }
    }
    with patch("autodetect_sources.check_gh_auth", return_value=True), \
         patch("autodetect_sources.check_s3_prefix", return_value=False), \
         patch("collect_github.fetch_repo_data", return_value=empty_raw):
        output = ra.cmd_run(cfg_path)

    assert Path(output).exists()
    assert "no-data" in str(output).lower()
