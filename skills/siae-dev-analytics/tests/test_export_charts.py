"""Test export_charts.py — openpyxl chart factory."""
from openpyxl import Workbook

import export_charts as ec


def _make_wb_with_data_sheet(n_devs: int = 3) -> Workbook:
    """Helper: crea wb con sheet 'Per Developer' con dati finti."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Per Developer"
    headers = ["developer"] + [f"kpi_{i}" for i in range(1, 15)]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r in range(2, n_devs + 2):
        ws.cell(row=r, column=1, value=f"dev_{r-1}")
        for c in range(2, 15):
            ws.cell(row=r, column=c, value=(r - 1) * c * 0.1)
    return wb


def test_add_velocity_bar_chart_creates_chart():
    wb = _make_wb_with_data_sheet(3)
    ws = wb.create_sheet("Trends")
    ec.add_velocity_bar_chart(ws, "Per Developer", 3)
    assert len(ws._charts) == 1


def test_add_velocity_bar_chart_skips_zero_devs():
    wb = Workbook()
    ws = wb.active
    ec.add_velocity_bar_chart(ws, "Sheet", 0)
    assert len(ws._charts) == 0


def test_add_velocity_quality_scatter_creates_chart():
    wb = _make_wb_with_data_sheet(3)
    ws = wb.create_sheet("Trends")
    ec.add_velocity_quality_scatter(ws, "Per Developer", 3)
    assert len(ws._charts) == 1


def test_add_before_after_bar_chart_creates_chart():
    wb = Workbook()
    ws = wb.active
    ec.add_before_after_bar_chart(ws, ["kpi_a", "kpi_b"], [10.0, 20.0], [15.0, 18.0])
    assert len(ws._charts) == 1


def test_add_before_after_empty_skips():
    wb = Workbook()
    ws = wb.active
    ec.add_before_after_bar_chart(ws, [], [], [])
    assert len(ws._charts) == 0


def test_add_pr_types_pie_creates_chart():
    wb = Workbook()
    ws = wb.active
    ec.add_pr_types_pie_chart(ws, {"feat": 10, "fix": 5, "refactor": 3})
    assert len(ws._charts) == 1


def test_add_ai_vs_manual_pie_creates_chart():
    wb = Workbook()
    ws = wb.active
    ec.add_ai_vs_manual_pie(ws, 0.65)
    assert len(ws._charts) == 1
    # Check that manual = 0.35
    assert ws.cell(row=101, column=2).value == 0.35
