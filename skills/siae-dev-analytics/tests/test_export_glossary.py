"""Test export_glossary.py — sheet Glossario KPI."""
from pathlib import Path

import export_glossary as eg


def test_load_glossary_data_returns_list():
    glossary = eg.load_glossary_data(
        Path(__file__).parent.parent / "reference" / "kpi-glossary-data.yaml"
    )
    assert isinstance(glossary, list)
    assert len(glossary) >= 50  # 53 KPI base; 68 include computed deltas


def test_load_glossary_missing_file_returns_empty():
    assert eg.load_glossary_data(Path("/nonexistent/path.yaml")) == []


def test_write_glossary_sheet_creates_tab(tmp_path):
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    eg.write_glossary_sheet(wb, [{"id": "test_kpi", "label_it": "Test", "formula": "x",
                                  "interpretation": "y", "good_if": "z", "benchmark": "w"}])
    out = tmp_path / "t.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    assert "Glossario KPI" in wb2.sheetnames


def test_glossary_sheet_has_6_columns(tmp_path):
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    eg.write_glossary_sheet(wb, [])
    out = tmp_path / "t.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["Glossario KPI"]
    headers = [c.value for c in ws[3]]
    assert len([h for h in headers if h]) == 6
    assert "Formula" in headers


def test_glossary_malformed_yaml_returns_empty(tmp_path):
    bad = tmp_path / "bad.yaml"
    bad.write_text("not yaml: {[}")
    assert eg.load_glossary_data(bad) == []
