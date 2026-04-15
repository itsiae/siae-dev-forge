"""Test per export_excel.py."""
from __future__ import annotations

from pathlib import Path
import pandas as pd
import pytest
from openpyxl import load_workbook

import export_excel as ee


@pytest.fixture
def sample_kpis_df() -> pd.DataFrame:
    """DataFrame con 3 dev + 14 colonne (11 KPI + 3 score)."""
    data = {
        "pr_cycle_time_p50": {"alice": 30.0, "bob": 7.0, "carol": 20.0},
        "lead_time_to_merge_p50": {"alice": 39.5, "bob": 13.5, "carol": 21.0},
        "pr_throughput_weekly": {"alice": 0.47, "bob": 0.47, "carol": 0.23},
        "time_to_first_review_p50": {"alice": 13.0, "bob": 1.5, "carol": 2.0},
        "deploy_frequency_monthly": {"alice": 1.0, "bob": 1.0, "carol": 0.0},
        "review_comments_p50": {"alice": 3.5, "bob": 1.5, "carol": 1.0},
        "rework_ratio": {"alice": 0.0, "bob": 0.0, "carol": 0.0},
        "test_presence_rate": {"alice": 0.5, "bob": 0.5, "carol": 1.0},
        "verification_rate": {"alice": 0.5, "bob": 0.5, "carol": 0.0},
        "design_driven_rate": {"alice": 0.5, "bob": 0.5, "carol": 0.0},
        "revert_rate": {"alice": 0.0, "bob": 0.5, "carol": 0.0},
        "velocity_score": {"alice": -0.2, "bob": 0.5, "carol": -0.3},
        "quality_score": {"alice": 0.3, "bob": -0.5, "carol": 0.2},
        "roi_index": {"alice": -0.06, "bob": -0.25, "carol": -0.06},
    }
    return pd.DataFrame(data)


@pytest.fixture
def sample_prs_df() -> pd.DataFrame:
    return pd.DataFrame([
        {"repo": "itsiae/r", "number": 1, "author": "alice", "cycle_time_hours": 4.0},
        {"repo": "itsiae/r", "number": 2, "author": "bob", "cycle_time_hours": 6.0},
    ])


@pytest.fixture
def sample_source_report() -> dict:
    return {
        "github": True,
        "s3_devforge": False,
        "s3_blend": False,
        "mode": "GITHUB-ONLY",
    }


def test_export_creates_file(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Export produce file xlsx."""
    out = tmp_path / "report.xlsx"
    ee.export(
        kpis_df=sample_kpis_df,
        raw_prs=sample_prs_df,
        source_report=sample_source_report,
        window=("2026-03-01", "2026-03-31"),
        output_path=out,
    )
    assert out.exists()
    assert out.stat().st_size > 0


def test_export_has_5_sheets_with_executive_summary(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Xlsx contiene 5 sheet: Executive Summary prima di tutti."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Executive Summary", "Summary", "Per Developer", "Raw Data", "Data Sources"}
    # Executive Summary deve essere il PRIMO sheet (tab leftmost)
    assert wb.sheetnames[0] == "Executive Summary"


def test_executive_summary_has_health_score(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Executive Summary contiene Team Health Score 0-100 + tier label."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Executive Summary"].iter_rows() for c in row if c.value)
    assert "Health" in text or "Salute" in text
    # Uno dei tier label deve essere presente
    assert any(t in text for t in ["Eccellente", "Buono", "Adeguato", "Da migliorare", "Critico", "N/A"])


def test_executive_summary_has_dora_tier(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Executive Summary dichiara DORA tier (Elite/High/Medium/Low)."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Executive Summary"].iter_rows() for c in row if c.value)
    assert "DORA" in text
    assert any(t in text for t in ["Elite", "High", "Medium", "Low"])


def test_executive_summary_has_data_quality_warnings(tmp_path, sample_source_report):
    """Se deploy_frequency=0 per tutti → warning SPECIFICO su tag convention SIAE."""
    # Tutti i KPI a 0 → trigger warnings specifici
    df = pd.DataFrame({
        "pr_cycle_time_p50": {"a": 10.0},
        "lead_time_to_merge_p50": {"a": 15.0},
        "pr_throughput_weekly": {"a": 1.0},
        "time_to_first_review_p50": {"a": 2.0},
        "deploy_frequency_monthly": {"a": 0.0},
        "review_comments_p50": {"a": 1.0},
        "rework_ratio": {"a": 0.0},
        "test_presence_rate": {"a": 0.5},
        "verification_rate": {"a": 0.0},
        "design_driven_rate": {"a": 0.0},
        "revert_rate": {"a": 0.0},
        "velocity_score": {"a": 0.0},
        "quality_score": {"a": 0.0},
        "roi_index": {"a": 0.0},
    })
    out = tmp_path / "report.xlsx"
    ee.export(df, pd.DataFrame(), sample_source_report, ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Executive Summary"].iter_rows() for c in row if c.value)
    # Assertion specifica: il messaggio deploy warning DEVE apparire
    assert "tag convention SIAE" in text, f"Missing specific deploy warning. Text: {text[:500]}"
    assert "verified-by" in text, f"Missing verification warning. Text: {text[:500]}"


def test_health_score_insufficient_sample_with_n_lt_3(tmp_path, sample_source_report):
    """N<3 dev → health score dichiara 'campione insufficiente' (z-score instabile)."""
    df = pd.DataFrame({
        "pr_cycle_time_p50": {"a": 5.0, "b": 10.0},
        "velocity_score": {"a": 0.5, "b": -0.5},
        "quality_score": {"a": 0.3, "b": -0.3},
        "roi_index": {"a": 0.15, "b": 0.15},
    })
    out = tmp_path / "report.xlsx"
    ee.export(df, pd.DataFrame(), sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Executive Summary"].iter_rows() for c in row if c.value)
    assert "campione insufficiente" in text


def test_dora_tier_zero_cycle_time_is_elite_not_na(tmp_path, sample_source_report):
    """Cycle time 0.0 (instant merge) → Elite, non N/A."""
    df = pd.DataFrame({
        "pr_cycle_time_p50": {"a": 0.0, "b": 0.0, "c": 0.0},
        "velocity_score": {"a": 0.5, "b": 0.5, "c": 0.5},
        "quality_score": {"a": 0.5, "b": 0.5, "c": 0.5},
        "roi_index": {"a": 0.25, "b": 0.25, "c": 0.25},
    })
    out = tmp_path / "report.xlsx"
    ee.export(df, pd.DataFrame(), sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Executive Summary"].iter_rows() for c in row if c.value)
    # Cycle 0 deve essere Elite (auto-merge è legittimo)
    assert "Elite" in text


def test_per_developer_has_italian_labels(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Per Developer header usa label italiani executive-friendly."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    header = [str(c.value) for c in wb["Per Developer"][1]]
    # Almeno 3 label italiane attese
    italian_labels = ["Tempo mediano PR (h)", "PR per settimana", "% PR con test", "Indice ROI"]
    found = sum(1 for lbl in italian_labels if lbl in header)
    assert found >= 3, f"Mancano label italiane. Header: {header}"


def test_summary_has_mode_and_window(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Summary contiene mode e finestra."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    summary = wb["Summary"]
    text = " ".join(str(c.value) for row in summary.iter_rows() for c in row if c.value)
    assert "GITHUB-ONLY" in text
    assert "2026-03-01" in text
    assert "2026-03-31" in text


def test_summary_has_confidential_header(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Summary ha header CONFIDENZIALE."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value)
    assert "CONFIDENZIALE" in text.upper()


def test_per_developer_has_all_kpi_columns(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Sheet Per Developer ha 14 colonne KPI + 1 col dev (con label italiani)."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    header = [c.value for c in wb["Per Developer"][1]]
    assert "Sviluppatore" in header
    # Label italiani per KPI principali
    for label in ["Tempo mediano PR (h)", "Indice velocità", "Indice qualità", "Indice ROI"]:
        assert label in header, f"Missing label: {label}"


def test_data_sources_sheet_declares_mode(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Data Sources elenca quali fonti sono usate."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    text = " ".join(str(c.value) for row in wb["Data Sources"].iter_rows() for c in row if c.value)
    assert "github" in text.lower()
    assert "s3" in text.lower() or "telemetry" in text.lower()
    assert "GITHUB-ONLY" in text


def test_raw_data_contains_prs(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Sheet Raw Data contiene PR raw."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out)
    wb = load_workbook(out)
    ws = wb["Raw Data"]
    # header + 2 righe dati
    assert ws.max_row >= 3


def test_anonymize_replaces_logins(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """anonymize=True → nessun login originale nel file."""
    out = tmp_path / "report.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out, anonymize=True)
    wb = load_workbook(out)
    all_text = ""
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for cell in row:
                if cell.value:
                    all_text += str(cell.value) + " "
    assert "alice" not in all_text
    assert "bob" not in all_text
    assert "carol" not in all_text


def test_reproducibility_same_input_same_checksum(tmp_path, sample_kpis_df, sample_prs_df, sample_source_report):
    """Stesso input → stesso file (escludendo metadata timestamp)."""
    import hashlib
    out1 = tmp_path / "r1.xlsx"
    out2 = tmp_path / "r2.xlsx"
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out1, generated_at="2026-04-15T00:00:00Z")
    ee.export(sample_kpis_df, sample_prs_df, sample_source_report,
              ("2026-03-01", "2026-03-31"), out2, generated_at="2026-04-15T00:00:00Z")
    # Openpyxl include metadata auto — testiamo che le cell values coincidano
    from openpyxl import load_workbook
    wb1 = load_workbook(out1)
    wb2 = load_workbook(out2)
    for sheet in wb1.sheetnames:
        vals1 = [c.value for row in wb1[sheet].iter_rows() for c in row]
        vals2 = [c.value for row in wb2[sheet].iter_rows() for c in row]
        assert vals1 == vals2, f"sheet {sheet} differs"
